#!/usr/bin/env python3
"""
GSC Analyser for ITR Stats
Usage: python3 scripts/gsc.py <filename>
       python3 scripts/gsc.py "~/Downloads/itrstats.in-Performance-on-Search-2026-03-07.xlsx"
       python3 scripts/gsc.py  (will prompt for file)

Saves report to analytics/gsc-YYYY-MM-DD.md
"""

import sys
import os
from datetime import date
from pathlib import Path

BRAND_TERMS = ['itrstats', 'itr stats', 'itrstats.in', 'itr stat']
REPO_ROOT = Path(__file__).parent.parent
ANALYTICS_DIR = REPO_ROOT / 'analytics'

# ── Helpers ──────────────────────────────────────────────────────────────────

def pct(n, d):
    return n / d if d else 0

def fmt_pct(f):
    return f"{f:.1%}"

def fmt_num(n):
    return f"{n:,.0f}"

def is_brand(query):
    q = str(query).lower()
    return any(t in q for t in BRAND_TERMS)

def md_table(headers, rows):
    sep = '|' + '|'.join(['---'] * len(headers)) + '|'
    head = '| ' + ' | '.join(headers) + ' |'
    body = '\n'.join('| ' + ' | '.join(str(c) for c in row) + ' |' for row in rows)
    return f"{head}\n{sep}\n{body}"

# ── Parsing ───────────────────────────────────────────────────────────────────

def load_sheet(wb, name):
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] is not None]
    return rows

def resolve_path(arg):
    p = Path(arg).expanduser()
    if p.is_absolute():
        return p
    # Try Downloads
    dl = Path.home() / 'Downloads' / arg
    if dl.exists():
        return dl
    raise FileNotFoundError(f"Cannot find file: {arg}\nTried: {p} and {dl}")

# ── Analysis ──────────────────────────────────────────────────────────────────

def analyse(filepath):
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not found. Installing...")
        os.system(f"{sys.executable} -m pip install openpyxl -q")
        import openpyxl

    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(filepath)

    # ── Raw data ──
    chart   = load_sheet(wb, 'Chart')      # Date, Clicks, Impressions, CTR, Position
    queries = load_sheet(wb, 'Queries')    # Query, Clicks, Impressions, CTR, Position
    pages   = load_sheet(wb, 'Pages')      # Page, Clicks, Impressions, CTR, Position
    countries = load_sheet(wb, 'Countries')
    devices = load_sheet(wb, 'Devices')
    filters = load_sheet(wb, 'Filters') if 'Filters' in wb.sheetnames else []

    # ── Filters ──
    date_range_raw = 'Unknown'
    for f in filters:
        if str(f[0]).lower() == 'date':
            date_range_raw = f[1]

    # ── Overall totals ──
    total_clicks = sum(r[1] for r in chart)
    total_imp    = sum(r[2] for r in chart)
    overall_ctr  = pct(total_clicks, total_imp)
    date_start   = chart[0][0] if chart else 'n/a'
    date_end     = chart[-1][0] if chart else 'n/a'

    # ── Trend: split halves ──
    mid = len(chart) // 2
    p1, p2 = chart[:mid], chart[mid:]
    p1_clicks = sum(r[1] for r in p1)
    p1_imp    = sum(r[2] for r in p1)
    p2_clicks = sum(r[1] for r in p2)
    p2_imp    = sum(r[2] for r in p2)
    p1_ctr    = pct(p1_clicks, p1_imp)
    p2_ctr    = pct(p2_clicks, p2_imp)
    p1_pos    = sum(r[4] for r in p1) / len(p1) if p1 else 0
    p2_pos    = sum(r[4] for r in p2) / len(p2) if p2 else 0
    ctr_chg   = pct(p2_ctr - p1_ctr, p1_ctr) if p1_ctr else 0
    imp_chg   = pct(p2_imp - p1_imp, p1_imp) if p1_imp else 0

    # Trend label
    if imp_chg > 0.5 and ctr_chg < -0.2:
        trend_label = "IMPRESSIONS SURGING, CTR FALLING — new pages indexed, CTR not yet optimised"
        trend_note  = "New content is getting indexed. Pages are appearing but titles/metas need work."
    elif p2_clicks < p1_clicks and p2_imp < p1_imp:
        trend_label = "CLICKS AND IMPRESSIONS FALLING — ranking drops or seasonality"
        trend_note  = "Check if any pages were recently modified. Could also be seasonal dip."
    elif abs(imp_chg) < 0.1 and ctr_chg > 0.1:
        trend_label = "STABLE IMPRESSIONS, CTR RISING — optimisations working"
        trend_note  = "Good signal. Continue the same approach."
    else:
        trend_label = "MIXED / EARLY STAGE — insufficient data for clear pattern"
        trend_note  = "Site is too new for a clear trend. Watch week-on-week from here."

    # ── Brand vs Non-brand ──
    brand_clicks = sum(r[1] for r in queries if is_brand(r[0]))
    brand_imp    = sum(r[2] for r in queries if is_brand(r[0]))
    nb_clicks    = sum(r[1] for r in queries if not is_brand(r[0]))
    nb_imp       = sum(r[2] for r in queries if not is_brand(r[0]))
    total_q_clicks = brand_clicks + nb_clicks

    # ── Devices ──
    total_dev_clicks = sum(r[1] for r in devices) or 1

    # ── Countries ──
    total_c_clicks = sum(r[1] for r in countries) or 1
    total_c_imp    = sum(r[2] for r in countries) or 1
    india_row = next((r for r in countries if str(r[0]).lower() == 'india'), None)
    us_row    = next((r for r in countries if 'united states' in str(r[0]).lower()), None)

    # ── Pages: opportunities ──
    high_imp_low_ctr = []  # impressions >= 50, CTR < 5%
    zero_click_pages = []  # 0 clicks, >= 10 impressions

    for p in pages:
        url, clicks, imp, ctr, pos = p[0], p[1] or 0, p[2] or 0, p[3] or 0, p[4] or 0
        gap = imp * 0.05 - clicks
        critical = pos <= 5 and ctr < 0.05
        if imp >= 50 and ctr < 0.05:
            high_imp_low_ctr.append((url, clicks, imp, ctr, pos, gap, critical))
        if clicks == 0 and imp >= 10:
            zero_click_pages.append((url, imp, pos))

    high_imp_low_ctr.sort(key=lambda x: -x[2])  # sort by impressions desc
    zero_click_pages.sort(key=lambda x: -x[1])

    # ── Queries: opportunities ──
    high_imp_low_ctr_q = []  # impressions >= 5, CTR < 5%
    near_top_q         = []  # position 4-10, >= 3 impressions

    for q in queries:
        if is_brand(q[0]):
            continue
        qname, clicks, imp, ctr, pos = q[0], q[1] or 0, q[2] or 0, q[3] or 0, q[4] or 0
        if imp >= 5 and ctr < 0.05:
            critical = pos <= 5 and clicks == 0
            high_imp_low_ctr_q.append((qname, clicks, imp, ctr, pos, critical))
        if 4 <= pos <= 10 and imp >= 3:
            near_top_q.append((qname, imp, pos, round(imp * 0.10)))

    high_imp_low_ctr_q.sort(key=lambda x: -x[2])
    near_top_q.sort(key=lambda x: -x[1])

    # ── Build report ──────────────────────────────────────────────────────────

    lines = []
    A = lines.append  # shorthand

    filename = Path(filepath).name
    today = date.today().strftime('%Y-%m-%d')

    A(f"# GSC Analysis — itrstats.in")
    A(f"**Export file:** {filename}")
    A(f"**Filter:** {date_range_raw}")
    A(f"**Period covered:** {date_start} to {date_end}")
    A(f"**Report date:** {today}")
    A(f"**Scope:** All pages, all queries — full site")
    A("")
    A("---")
    A("")

    # Core metrics
    A("## Core Metrics")
    A("")
    A(md_table(
        ['Metric', 'Value'],
        [
            ['Total clicks', fmt_num(total_clicks)],
            ['Total impressions', fmt_num(total_imp)],
            ['Overall CTR', fmt_pct(overall_ctr)],
            ['Date range', f"{date_start} to {date_end}"],
            ['Unique queries tracked', fmt_num(len(queries))],
            ['Pages tracked', fmt_num(len(pages))],
        ]
    ))
    A("")
    A("> Note: Overall CTR is inflated by brand traffic. Non-brand CTR is far lower.")
    A("")
    A("---")
    A("")

    # Brand vs non-brand
    A("## Brand vs Non-Brand")
    A("")
    A(md_table(
        ['Category', 'Clicks', 'Impressions', 'Click Share'],
        [
            ['Brand (itrstats, itr stats, itrstats.in)',
             fmt_num(brand_clicks), fmt_num(brand_imp),
             fmt_pct(pct(brand_clicks, total_q_clicks))],
            ['Non-brand',
             fmt_num(nb_clicks), fmt_num(nb_imp),
             fmt_pct(pct(nb_clicks, total_q_clicks))],
        ]
    ))
    A("")
    if total_q_clicks > 0:
        nb_share = pct(nb_clicks, total_q_clicks)
        if nb_share < 0.05:
            A(f"**{fmt_pct(pct(brand_clicks, total_q_clicks))} of clicks are brand.** Non-brand = {fmt_num(nb_clicks)} clicks over this period.")
            A("Non-brand impressions are building — pages are showing up but not yet getting clicked. Fix titles/metas.")
        elif nb_share < 0.20:
            A(f"Non-brand share is growing ({fmt_pct(nb_share)} of clicks). Early traction visible.")
        else:
            A(f"Healthy non-brand share at {fmt_pct(nb_share)}. Content SEO is working.")
    A("")
    A("---")
    A("")

    # Trend
    A("## Trend Analysis")
    A("")
    A(f"**Pattern: {trend_label}**")
    A("")
    A(md_table(
        ['Period', 'Dates', 'Clicks', 'Impressions', 'CTR', 'Avg Position'],
        [
            ['Period 1 (early)', f"{p1[0][0]} to {p1[-1][0]}" if p1 else 'n/a',
             fmt_num(p1_clicks), fmt_num(p1_imp), fmt_pct(p1_ctr), f"{p1_pos:.2f}"],
            ['Period 2 (recent)', f"{p2[0][0]} to {p2[-1][0]}" if p2 else 'n/a',
             fmt_num(p2_clicks), fmt_num(p2_imp), fmt_pct(p2_ctr), f"{p2_pos:.2f}"],
            ['Change', '',
             f"{pct(p2_clicks-p1_clicks, p1_clicks):+.1%}" if p1_clicks else 'n/a',
             f"{imp_chg:+.1%}", f"{ctr_chg:+.1%}",
             f"{p2_pos-p1_pos:+.2f}"],
        ]
    ))
    A("")
    A(trend_note)
    A("")
    A("---")
    A("")

    # Geographic
    A("## Geographic Breakdown (Top 10 Countries)")
    A("")
    country_rows = []
    for c in countries[:10]:
        country_rows.append([
            c[0],
            fmt_num(c[1]),
            fmt_pct(pct(c[1], total_c_clicks)),
            fmt_num(c[2]),
            fmt_pct(pct(c[2], total_c_imp)),
            fmt_pct(c[3] or 0),
            f"{c[4]:.1f}" if c[4] else 'n/a'
        ])
    A(md_table(['Country', 'Clicks', '% Clicks', 'Impressions', '% Impressions', 'CTR', 'Position'], country_rows))
    A("")
    if us_row:
        us_clicks, us_imp, us_ctr = us_row[1] or 0, us_row[2] or 0, us_row[3] or 0
        if us_imp >= 100 and us_ctr < 0.02:
            A(f"**US has {fmt_num(us_imp)} impressions but {fmt_pct(us_ctr)} CTR** — income data pages attracting international researchers who aren't clicking. Titles need to address this audience.")
    A("")
    A("---")
    A("")

    # Devices
    A("## Device Split")
    A("")
    dev_rows = [[d[0], fmt_num(d[1]), fmt_pct(pct(d[1], total_dev_clicks)), fmt_pct(d[3] or 0)] for d in devices]
    A(md_table(['Device', 'Clicks', '% of Clicks', 'CTR'], dev_rows))
    A("")
    A("---")
    A("")

    # Opportunities — pages
    A("## Opportunities")
    A("")
    A("### High-Impression / Low-CTR Pages (impressions >= 50, CTR < 5%)")
    A("")
    if high_imp_low_ctr:
        page_rows = []
        for url, clicks, imp, ctr, pos, gap, critical in high_imp_low_ctr:
            flag = " **CRITICAL**" if critical else ""
            page_rows.append([
                url.replace('https://itrstats.in', ''),
                fmt_num(clicks),
                fmt_num(imp),
                fmt_pct(ctr),
                f"{pos:.1f}",
                f"+{gap:.0f}",
                flag.strip() or '-'
            ])
        A(md_table(['Page', 'Clicks', 'Impressions', 'CTR', 'Position', 'Potential Gap', 'Flag'], page_rows))
    else:
        A("No pages with >= 50 impressions and CTR < 5%.")
    A("")

    # 0-click pages
    A("### Pages with 0 Clicks and >= 10 Impressions")
    A("")
    if zero_click_pages:
        A(md_table(
            ['Page', 'Impressions', 'Position'],
            [[url.replace('https://itrstats.in', ''), fmt_num(imp), f"{pos:.1f}"]
             for url, imp, pos in zero_click_pages]
        ))
    else:
        A("None.")
    A("")

    # High-impression / low-CTR queries
    A("### High-Impression / Low-CTR Queries (impressions >= 5, CTR < 5%, non-brand)")
    A("")
    if high_imp_low_ctr_q:
        q_rows = []
        for qname, clicks, imp, ctr, pos, critical in high_imp_low_ctr_q[:20]:
            flag = "CRITICAL" if critical else "-"
            q_rows.append([qname, fmt_num(clicks), fmt_num(imp), fmt_pct(ctr), f"{pos:.1f}", flag])
        A(md_table(['Query', 'Clicks', 'Impressions', 'CTR', 'Position', 'Flag'], q_rows))
    else:
        A("No qualifying queries.")
    A("")

    # Near-top queries
    A("### Near-the-Top Queries (position 4-10, >= 3 impressions)")
    A("")
    if near_top_q:
        A(md_table(
            ['Query', 'Impressions', 'Position', 'Target clicks at pos 3'],
            [[q, fmt_num(imp), f"{pos:.1f}", f"~{est}"] for q, imp, pos, est in near_top_q[:15]]
        ))
    else:
        A("No qualifying queries.")
    A("")
    A("---")
    A("")

    # Priority actions
    A("## Priority Action List")
    A("")
    A("### P0 — This Week (CTR fixes, no new content)")
    A("")

    p0_items = []
    # Biggest page gaps
    for url, clicks, imp, ctr, pos, gap, critical in high_imp_low_ctr[:5]:
        slug = url.replace('https://itrstats.in', '') or '/'
        p0_items.append(f"**{slug}** — {fmt_num(imp)} impressions, {fmt_pct(ctr)} CTR, pos {pos:.1f}. Gap: +{gap:.0f} clicks. Rewrite title and meta description.")

    # Critical position-1-3 pages with near-zero CTR
    for p in pages:
        url, clicks, imp, ctr, pos = p[0], p[1] or 0, p[2] or 0, p[3] or 0, p[4] or 0
        if pos <= 3 and imp >= 10 and ctr < 0.05:
            slug = url.replace('https://itrstats.in', '') or '/'
            item = f"**{slug}** — CRITICAL: ranking position {pos:.1f} with {fmt_pct(ctr)} CTR. {fmt_num(imp)} impressions, {fmt_num(clicks)} clicks. Likely AI Overview consuming clicks — add a structured answer box at the top of the page."
            if item not in p0_items:
                p0_items.append(item)

    for i, item in enumerate(p0_items, 1):
        A(f"**P0.{i}** — {item}")
        A("")

    A("### P1 — Next Sprint (ranking improvements)")
    A("")
    p1_items = []
    # Near-top queries without dedicated page
    for q, imp, pos, est in near_top_q[:8]:
        p1_items.append(f"Query `{q}` — position {pos:.1f}, {fmt_num(imp)} impressions. Push to top 3 with internal links or content update on the ranking page.")
    for i, item in enumerate(p1_items, 1):
        A(f"**P1.{i}** — {item}")
        A("")

    A("### P2 — New Content (demand signals)")
    A("")
    # Queries with 0 clicks but >= 5 impressions (consistent demand)
    p2_queries = [(q, imp, pos) for q, clicks, imp, ctr, pos, _ in high_imp_low_ctr_q if clicks == 0 and imp >= 5]
    p2_queries.sort(key=lambda x: -x[1])
    for i, (q, imp, pos) in enumerate(p2_queries[:5], 1):
        A(f"**P2.{i}** — `{q}` — {fmt_num(imp)} impressions, position {pos:.1f}, 0 clicks. Consistent demand signal. Build dedicated page if Data Hook Score >= 5.")
        A("")

    A("---")
    A("")

    # Raw data
    A("## Raw Data")
    A("")
    A("### All Pages")
    A("")
    A(md_table(
        ['Page', 'Clicks', 'Impressions', 'CTR', 'Position'],
        [[p[0].replace('https://itrstats.in', '') or '/',
          fmt_num(p[1] or 0), fmt_num(p[2] or 0),
          fmt_pct(p[3] or 0), f"{p[4]:.2f}" if p[4] else 'n/a']
         for p in pages]
    ))
    A("")
    A("### All Queries (top 50 by impressions)")
    A("")
    q_sorted = sorted(queries, key=lambda x: -(x[2] or 0))
    A(md_table(
        ['Query', 'Clicks', 'Impressions', 'CTR', 'Position'],
        [[q[0], fmt_num(q[1] or 0), fmt_num(q[2] or 0),
          fmt_pct(q[3] or 0), f"{q[4]:.2f}" if q[4] else 'n/a']
         for q in q_sorted[:50]]
    ))

    report = '\n'.join(lines)
    return report, today

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) >= 2:
        arg = ' '.join(sys.argv[1:])  # handle spaces in filename
    else:
        arg = input("GSC export filename (or full path): ").strip()

    filepath = resolve_path(arg)
    print(f"Parsing: {filepath}")

    report, today = analyse(filepath)

    ANALYTICS_DIR.mkdir(exist_ok=True)
    out_path = ANALYTICS_DIR / f"gsc-{today}.md"
    out_path.write_text(report)
    print(f"\nReport saved: {out_path}")

    # Print summary stats from report (first ~30 lines)
    preview = '\n'.join(report.split('\n')[:35])
    print(f"\n--- Preview ---\n{preview}\n...")

if __name__ == '__main__':
    main()
